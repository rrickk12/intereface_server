import csv
import openpyxl
from datetime import datetime
import os

csv_path = os.path.join('static', 'downloads', 'temp_odoo.csv')
os.makedirs(os.path.dirname(csv_path), exist_ok=True)


import csv
import openpyxl
import os


def fill_procfy_template(novo_csv_path, procfy_template_path, output_xlsx_path):
    """
    Lê um CSV no formato:
      Data, Descrição, Valor, Tipo Transação, Documento, Contato, Tipo, Categoria, Memo, Status
    e preenche o template XLSX do Procfy com o seguinte mapeamento:
    
      - Tipo de Lançamento         = valor da coluna "Tipo"
      - Data de pagamento           = valor da coluna "Data"
      - Data de competência        = (em branco)
      - Descrição                  = "Memo" (se vazio, usa "Descrição")
      - Valor                      = valor absoluto do campo "Valor" com 2 casas decimais
      - Categoria                  = valor da coluna "Categoria"
      - Recebido de/Pago a         = valor da coluna "Contato"
      - Pago?                      = "Sim"
      - Observações                = "Preenchido Automaticamente"
      - Contas bancárias           = "Sicoob 2"
      - Número de Documento / Nosso número = (em branco)
      - Modo de pagamento          = (em branco)
      - Centro de Custo            = "Indefinido"
      - Tags (Por Marcação)        = (em branco)
    """
    # Carrega o template XLSX do Procfy
    workbook = openpyxl.load_workbook(procfy_template_path)
    sheet = workbook.active  # ou workbook['NomeDaAba'] se necessário

    # Lê o cabeçalho do template
    procfy_headers = [cell.value for cell in sheet[1]]

    # Lê o CSV com detecção automática do delimitador
    with open(novo_csv_path, 'r', encoding='utf-8') as csv_file:
        sample = csv_file.read(1024)
        csv_file.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ','
        print("Delimitador detectado:", dialect.delimiter)  # Debug
        reader = csv.DictReader(csv_file, dialect=dialect)
        csv_rows = list(reader)
        print("Número de linhas lidas do CSV:", len(csv_rows))  # Debug

    current_row = 2
    for row in csv_rows:
        # Se "Memo" estiver vazio, usa "Descrição"
        descricao = row.get("Memo", "").strip() or row.get("Descrição", "").strip()
        try:
            # Trata valores com vírgula, converte para float e aplica valor absoluto
            valor_str = row.get("Valor", "0").strip().replace(",", ".")
            valor = abs(float(valor_str))
        except ValueError:
            valor = 0.0

        # Caso queira que o valor seja escrito como número no Excel com 2 casas decimais:
        # row_values["Valor"] conterá o valor numérico e usaremos cell.number_format.
        row_values = {
            "Tipo de Lançamento": row.get("Tipo", "").strip(),
            "Data de pagamento": row.get("Data", "").strip(),
            "Data de competência": "",
            "Descrição": descricao,
            "Valor": valor,  # valor numérico
            "Categoria": row.get("Categoria", "").strip(),
            "Recebido de/Pago a": row.get("Contato", "").strip(),
            "Pago?": "Sim",
            "Observações": "Preenchido Automaticamente",
            "Contas bancárias": "Sicoob 2",
            "Número de Documento / Nosso número": "",
            "Modo de pagamento": "",
            "Centro de Custo": "Indefinido",
            "Tags (Por Marcação)": ""
        }

        # Preenche a linha no template respeitando a ordem do cabeçalho
        for col_index, header_name in enumerate(procfy_headers, start=1):
            cell = sheet.cell(row=current_row, column=col_index)
            if header_name == "Valor":
                # Escreve o valor numérico e define o formato para 2 casas decimais
                cell.value = row_values.get("Valor", 0.0)
                cell.number_format = '0.00'
                # Se preferir armazenar como string com vírgula, use:
                # cell.value = f"{valor:.2f}".replace(".", ",")
            else:
                cell.value = row_values.get(header_name, "")
        current_row += 1

    os.makedirs(os.path.dirname(output_xlsx_path), exist_ok=True)
    workbook.save(output_xlsx_path)
    print("Arquivo XLSX salvo em:", output_xlsx_path)
    return output_xlsx_path

def process_excel_data():
    """
    Loads the Procfy XLSX template, fills in default values, and saves the result.
    """
    import openpyxl
    try:
        workbook = openpyxl.load_workbook("modelo_Procfy_MMG.xlsx")
        sheet = workbook.active
        # Read header from the first row
        cabecalho = [cell.value for cell in sheet[1]]
        # Default values to fill
        valores_padrao = {
            "Tipo de Lançamento": "Recebimentos",
            "Data de competência": "01/02/2025",
            "Descrição": "INV/",
            "Categoria": "Locação de equipamentos",
            "Recebido de/Pago a": "Cliente",
            "Pago?": "Não",
            "Modo de pagamento": "PIX"
        }
        # Map header names to column indices
        indices_colunas = {header: idx+1 for idx, header in enumerate(cabecalho)}
        for coluna, valor in valores_padrao.items():
            if coluna in indices_colunas:
                col_idx = indices_colunas[coluna]
                sheet.cell(row=2, column=col_idx).value = valor
        workbook.save("modelo_Procfy_MMG_preenchido.xlsx")
        return "Excel processed and saved as 'modelo_Procfy_MMG_preenchido.xlsx'."
    except Exception as e:
        return f"Error processing Excel: {e}"


def generate_procfy_xlsx(odoo_csv_path, procfy_xlsx_path):
    """
    Reads an Odoo invoices CSV and generates an XLSX formatted for Procfy import.
    
    odoo_csv_path: path to the CSV exported from Odoo.
    procfy_xlsx_path: output path for the XLSX file formatted for Procfy.
    """
    # Read the CSV
    with open(odoo_csv_path, 'r', encoding='utf-8') as f:
        csv_reader = csv.DictReader(f)
        rows = list(csv_reader)
    
    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Procfy header (order matters)
    header = [
        "Tipo de Lançamento",
        "Data de pagamento",
        "Data de competência",
        "Descrição",
        "Valor",
        "Categoria",
        "Recebido de/Pago a",
        "Pago?",
        "Observações",
        "Contas bancárias",
        "Número de Documento / Nosso número",
        "Modo de pagamento",
        "Centro de Custo",
        "Tags (Por Marcação)"
    ]
    sheet.append(header)

    # Iterate over each CSV row and map fields to Procfy columns
    for r in rows:
        tipo_lancamento = "Recebimentos" if r.get("move_type") == "out_invoice" else "Pagamentos"
        data_pagamento = r.get("invoice_date_due") or ""
        data_competencia = r.get("invoice_date") or ""
        descricao = r.get("name") or ""
        valor = r.get("amount_total") or ""
        categoria = "Locação de equipamentos"
        recebido_de = r.get("invoice_partner_display_name") or ""
        pago = "Sim" if r.get("payment_state") == "paid" else "Não"
        observacoes = ""
        contas_bancarias = ""
        numero_documento = r.get("name") or ""
        modo_pagamento = "PIX"
        centro_custo = ""
        tags = ""

        row_data = [
            tipo_lancamento,
            data_pagamento,
            data_competencia,
            descricao,
            valor,
            categoria,
            recebido_de,
            pago,
            observacoes,
            contas_bancarias,
            numero_documento,
            modo_pagamento,
            centro_custo,
            tags
        ]
        sheet.append(row_data)

    # Save the generated XLSX file
    workbook.save(procfy_xlsx_path)
    print(f"Arquivo gerado para o Procfy: {procfy_xlsx_path}")
